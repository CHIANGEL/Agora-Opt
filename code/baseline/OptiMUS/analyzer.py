import os
import math
from answer_and_dataset import get_desc_and_answer, get_answer_from_output
import pandas as pd
import re
import openpyxl
from openpyxl.styles import PatternFill


"""
INFO: analytical summaries
GRAPH: tables and plots
BASE: common utility analysis
OTHER: miscellaneous helpers
"""


def analyze_debate_converge(dataset_path):
    """
    Find debate-mode cases where initially different answers changed during the
    debate process.
    """
    for name in os.listdir(dataset_path):
        if not name.startswith("problem_"):
            continue
        i = int(name.split("_")[1])
        converge_txt_path = os.path.join(dataset_path, name, "log_converge.txt")
        
        with open(converge_txt_path, "r") as f:
            lines = f.readlines()
            if len(lines) < 2:
                continue
        
        first_line = lines[0].strip()
        last_line = lines[-1].strip()
        
        modelA_answer_first = first_line.split(",")[0].strip().split(":")[-1].strip()
        modelB_answer_first = first_line.split(",")[1].strip().split(":")[-1].strip()
        modelA_answer_last = last_line.split(",")[0].strip().split(":")[-1].strip()
        modelB_answer_last = last_line.split(",")[1].strip().split(":")[-1].strip()
        
        if modelA_answer_first != modelA_answer_last or modelB_answer_first != modelB_answer_last and modelA_answer_first != modelB_answer_first:
            print(f"Problem {i} converged: {modelA_answer_first} -> {modelA_answer_last}, {modelB_answer_first} -> {modelB_answer_last}")


def get_unfinished_problems(dataset_path):
    """
    Return a list of unfinished problem ids based on missing output files.
    """
    id_list = []
    for name in os.listdir(dataset_path):
        if not name.startswith("problem_"):
            continue
        idx = int(name.split("_")[1])
        output_file = os.path.join(dataset_path, name, "output_solution.txt")
        if not os.path.exists(output_file):
            id_list.append(idx)
            print(f"Problem {idx} has no output_solution.txt")
    return id_list


def get_model_results(dataset_path):
    """
    Return a dict of `{idx: (predicted_answer, ground_truth_answer)}`.
    """
    results = {}
    dataset_name = dataset_path.split("/")[-1]
    data = get_desc_and_answer(dataset_name)
    print("Ground truth answers loaded for", len(data), "problems.")
    for name in os.listdir(dataset_path):
        if not name.startswith("problem_"):
            continue
        idx = int(name.split("_")[1])
        output_file = os.path.join(dataset_path, name, "output_solution.txt")
        if not os.path.exists(output_file):
            print(f"Problem {idx} has no output_solution.txt, skipping.")
            my_answer = None
        else:
            with open(output_file, "r") as f:
                my_answer = f.read()
            try:
                my_answer = float(my_answer)
            except Exception as e:
                print(f"Problem {idx} has invalid answer format: my_answer={my_answer}")
                my_answer = None
        results[idx] = (my_answer, float(data[idx][1]))
    
    # Sort by problem index.
    results = dict(sorted(results.items(), key=lambda x: x[0]))
    return results

def compare_wrong_problems(model1_results, model2_results, tol=1e-1):
    """
    Return the wrong-problem sets for both models plus their overlaps and gaps.
    """
    wrong_model1 = set()
    wrong_model2 = set()
    for idx, (my_answer, gt_answer) in model1_results.items():
        if my_answer is None or type(my_answer) == str or math.fabs(my_answer - gt_answer) >= tol:
            wrong_model1.add(idx)
    for idx, (my_answer, gt_answer) in model2_results.items():
        if my_answer is None or type(my_answer) == str or math.fabs(my_answer - gt_answer) >= tol:
            wrong_model2.add(idx)
    intersection = wrong_model1.intersection(wrong_model2)
    diff_model1 = wrong_model1.difference(wrong_model2)
    diff_model2 = wrong_model2.difference(wrong_model1) 
    return wrong_model1, wrong_model2, intersection, diff_model1, diff_model2

def parse_converge_file(file_path):
    """
    Parse `log_converge.txt` and extract per-round values.
    """
    rounds_data = {0: {'A': '', 'B': ''}, 
                   1: {'A': '', 'B': ''}, 
                   2: {'A': '', 'B': ''}, 
                   3: {'A': '', 'B': ''}}
    
    try:
        with open(file_path, 'r') as f:
            for line in f:
                line = line.strip()
                # Match records such as: A0: None, B0: 10349200.0
                pattern = r'([AB])(\d+):\s*([^,]+)'
                matches = re.findall(pattern, line)
                
                for match in matches:
                    model, round_num, value = match
                    round_num = int(round_num)
                    
                    if value.strip() == 'None':
                        value = 'None'
                    
                    value = re.sub("`", "", value).strip()
                    if round_num in rounds_data:
                        rounds_data[round_num][model] = value
    except FileNotFoundError:
        print(f"Warning: file not found: {file_path}")
    except Exception as e:
        print(f"Error while parsing {file_path}: {e}")
    
    return rounds_data

def process_problem_directory(root_dir, dataset_name):
    """
    Process all problem directories and collect convergence data.
    """
    data = []
    
    # Load ground-truth answers for comparison.
    correct_answers = {}
    try:
        gt_data = get_desc_and_answer(dataset_name)
        for i, (desc, ans) in enumerate(gt_data):
            correct_answers[i] = float(ans)
    except Exception as e:
        print(f"Warning: failed to load ground-truth answers: {e}")
        correct_answers = {}
    
    def format_number(value):
        """Format a numeric value to one decimal place."""
        if not value or value == '' or value == 'None' or str(value).strip().lower() == 'none':
            return value
        try:
            num = float(value)
            return f"{num:.1f}"
        except (ValueError, TypeError):
            return value
    
    for i in range(len(correct_answers)):
        problem_dir = os.path.join(root_dir, f"problem_{i}")
        if not os.path.exists(problem_dir):
            print(f"Warning: directory does not exist: {problem_dir}")
            continue
            
        debate_file = os.path.join(problem_dir, "log_debate.txt")
        converge_file = os.path.join(problem_dir, "log_converge.txt")
        output_solution_file = os.path.join(problem_dir, "output_solution.txt")
        
        # Read the final answer.
        final_answer = ''
        if os.path.exists(output_solution_file):
            try:
                with open(output_solution_file, 'r') as f:
                    final_answer = f.read().strip()
                    final_answer = format_number(final_answer)
            except Exception as e:
                print(f"Error while reading {output_solution_file}: {e}")
                final_answer = ''
        
        # Skip problems without a final answer.
        if not final_answer or final_answer == '':
            print(f"Skipping problem {i}: missing final answer")
            continue
        
        # Check whether log_debate.txt recorded initial convergence.
        initial_converged = None
        if os.path.exists(debate_file):
            try:
                with open(debate_file, 'r') as f:
                    content = f.read()
                    match = re.search(r'Initial solutions converged:\s*([\d.]+)', content)
                    if match:
                        initial_converged = format_number(match.group(1))
            except Exception as e:
                print(f"Error while reading {debate_file}: {e}")
        
        # Get the ground-truth answer for this problem.
        gt_answer = correct_answers.get(i, None)
        
        if initial_converged:
            # If initial convergence exists, both models share the same round-0 value.
            data.append({
                'Problem': f'problem{i}',
                'Model': 'gpt-4o',
                'round0': initial_converged,
                'round1': '',
                'round2': '',
                'round3': '',
                'final_answer': final_answer,
                'gt_answer': gt_answer
            })
            data.append({
                'Problem': f'problem{i}',
                'Model': 'gemini-2.5-pro',
                'round0': initial_converged,
                'round1': '',
                'round2': '',
                'round3': '',
                'final_answer': final_answer,
                'gt_answer': gt_answer
            })
        else:
            # Otherwise parse log_converge.txt.
            rounds_data = parse_converge_file(converge_file)
            
            data.append({
                'Problem': f'problem{i}',
                'Model': 'gpt-4o',
                'round0': format_number(rounds_data[0]['A']),
                'round1': format_number(rounds_data[1]['A']),
                'round2': format_number(rounds_data[2]['A']),
                'round3': format_number(rounds_data[3]['A']),
                'final_answer': final_answer,
                'gt_answer': gt_answer
            })
            data.append({
                'Problem': f'problem{i}',
                'Model': 'gemini-2.5-pro',
                'round0': format_number(rounds_data[0]['B']),
                'round1': format_number(rounds_data[1]['B']),
                'round2': format_number(rounds_data[2]['B']),
                'round3': format_number(rounds_data[3]['B']),
                'final_answer': final_answer,
                'gt_answer': gt_answer
            })
    return data

def create_table(data, output_file="debate_converge_results.xlsx"):
    """
    Create a convergence table and optionally save it with color highlighting.
    """
    import openpyxl
    from openpyxl.styles import PatternFill
    import math
    
    df = pd.DataFrame(data)
    
    # Reorder columns and exclude gt_answer from the display table.
    df_display = df[['Problem', 'Model', 'round0', 'round1', 'round2', 'round3', 'final_answer']]
    
    if output_file:
        # Save to Excel and add color annotations.
        df_display.to_excel(output_file, index=False)
        
        # Reopen the file to add formatting.
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Define colors.
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # light green
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")    # light red
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")   # light gray
        
        # Color numeric cells row by row.
        for row_idx, row_data in enumerate(data, start=2):  # Row 1 is the header.
            gt_answer = row_data.get('gt_answer', None)
            if gt_answer is None:
                continue
                
            # Check round0-round3 and final_answer (columns C-G).
            for col_idx, round_col in enumerate(['round0', 'round1', 'round2', 'round3', 'final_answer'], start=3):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(round_col, '')
                
                if value and value != '':
                    if value == 'None' or str(value).strip().lower() == 'none':
                        # Mark None values in light gray.
                        cell.fill = gray_fill
                    else:
                        try:
                            # Values may already be formatted strings.
                            num_value = float(value)
                            # Compare against the ground-truth answer.
                            if abs(num_value - gt_answer) < 0.1:
                                cell.fill = green_fill
                            else:
                                cell.fill = red_fill
                        except (ValueError, TypeError):
                            # Leave non-numeric values uncolored.
                            pass
        
        wb.save(output_file)
        print(
            f"Results saved to {output_file} with color highlighting "
            "(green=correct, red=incorrect, gray=None)."
        )
    else:
        print(df_display.to_string(index=False))
    
    return df_display

def create_converge_table(directory, dataset_name, output="converge_results.xlsx"):
    """Create the convergence summary spreadsheet."""
    if not os.path.exists(directory):
        print(f"Error: directory does not exist: {directory}")
        return
    
    data = process_problem_directory(directory, dataset_name)
    
    if data:
        create_table(data, output)
        print("Finished processing.")
    else:
        print("No data found.")

def analyze_error_cases(path1, path2):
    """Analyze overlaps and differences among error cases."""
    model1_results = get_model_results(path1)
    model2_results = get_model_results(path2)
    wrong_model1, wrong_model2, intersection, diff_model1, diff_model2 = compare_wrong_problems(model1_results, model2_results)
    
    print("Model 1 wrong problems:", wrong_model1)
    print("Model 2 wrong problems:", wrong_model2)
    print("Intersection of wrong problems:", intersection)
    print("Model 1 unique wrong problems:", diff_model1)
    print("Model 2 unique wrong problems:", diff_model2)

def get_converge_length(debate_dataset_path):
    '''
    return {idx: converge_round} the number will be -1 if not converge finally, and will be 0 if converge at the beginning
    (INFO)
    '''
    converge_results = {}
    for name in os.listdir(debate_dataset_path):
        if not name.startswith("problem_"):
            continue
        idx = int(name.split("_")[1])
        converge_txt_path = os.path.join(debate_dataset_path, name, "log_converge.txt")
        # If the file is empty, treat it as initial convergence.
        with open(converge_txt_path, "r") as f:
            lines = f.readlines()
            if len(lines) == 0:
                converge_results[idx] = 0
                continue
            else:
                last_line = lines[-1].strip()
                length = len(lines)
                model_A_answer_last = last_line.split(",")[0].strip().split(":")[-1].strip()
                model_B_answer_last = last_line.split(",")[1].strip().split(":")[-1].strip()
            if model_A_answer_last == 'None' or model_B_answer_last == 'None':
                converge_results[idx] = -1
                continue
            try:
                delta = math.fabs(float(model_A_answer_last) - float(model_B_answer_last))
                if delta < 0.1:
                    converge_results[idx] = length - 1
                else:
                    converge_results[idx] = -1
            except Exception as e:
                print(f"Problem {idx} has invalid answer format: A={model_A_answer_last}, B={model_B_answer_last}")
                converge_results[idx] = -1
                continue
    
    # Compute the average converge length excluding non-converged cases.
    total_length = 0
    count = 0
    for idx, length in converge_results.items():
        if length > 0:
            total_length += length
            count += 1
    avg_length = total_length / count if count > 0 else 0
    print(f"Average converge length (Initial converge not included): {avg_length:.2f}")
    
    return converge_results
    
def analyze_converge_correct_rate(dataset_path):
    """Compute accuracy split by converged vs. non-converged cases."""
    model_results = get_model_results(dataset_path)
    converge_results = get_converge_length(dataset_path)
    
    # Measure accuracy among converged and non-converged subsets.
    total_converged = 0
    correct_converged = 0
    total_non_converged = 0
    correct_non_converged = 0
    for idx, (my_answer, gt_answer) in model_results.items():
        converge_round = converge_results.get(idx, -1)
        if converge_round >= 0:
            total_converged += 1
            if my_answer is not None and type(my_answer) != str and math.fabs(my_answer - gt_answer) < 0.1:
                correct_converged += 1
        else:
            total_non_converged += 1
            if my_answer is not None and type(my_answer) != str and math.fabs(my_answer - gt_answer) < 0.1:
                correct_non_converged += 1 
    
    print(f"Total problems: {len(model_results)}")
    print(f"Converged problems: {total_converged}, Correct: {correct_converged}, Accuracy: {correct_converged/total_converged if total_converged > 0 else 0:.2%}")
    print(f"Non-converged problems: {total_non_converged}, Correct: {correct_non_converged}, Accuracy: {correct_non_converged/total_non_converged if total_non_converged > 0 else 0:.2%}")
    print(f"Overall accuracy: {(correct_converged + correct_non_converged)/len(model_results):.2%}")
    return {
        "total_problems": len(model_results),
        "converged_problems": total_converged,
        "correct_converged": correct_converged,
        "accuracy_converged": correct_converged/total_converged if total_converged > 0 else 0,
        "non_converged_problems": total_non_converged,
        "correct_non_converged": correct_non_converged,
        "accuracy_non_converged": correct_non_converged/total_non_converged if total_non_converged > 0 else 0,
        "overall_accuracy": (correct_converged + correct_non_converged)/len(model_results)
    }
    
def pure_correct_rate(dataset_path):
    """Compute raw accuracy without convergence-specific analysis."""
    model_results = get_model_results(dataset_path)
    total = len(model_results)
    correct = 0
    for idx, (my_answer, gt_answer) in model_results.items():
        if my_answer is not None and type(my_answer) != str and math.fabs(my_answer - gt_answer) < 0.05:
            correct += 1
        else:
            print(f"Problem {idx} incorrect or invalid: my_answer={my_answer}, gt_answer={gt_answer}")
    print(f"Total problems: {total}, Correct: {correct}, Accuracy: {correct/total if total > 0 else 0:.2%}")
    return {
        "total_problems": total,
        "correct": correct,
        "accuracy": correct/total if total > 0 else 0
    }

    


if __name__ == "__main__":
    pure_correct_rate("history_single/gemini-2.5-pro_temperature_0.01/ComplexLP")

    
